import pandas as pd
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelDuplicatorApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel Product File Duplicator")
        master.geometry("700x580") # Increased size for better layout and log

        # Configure columns for better layout
        master.grid_columnconfigure(0, weight=0) # Labels column
        master.grid_columnconfigure(1, weight=1) # Entries column
        master.grid_columnconfigure(2, weight=0) # Buttons column

        # --- Variables to store paths ---
        self.lookup_file_path_var = tk.StringVar()
        self.source_folder_path_var = tk.StringVar()
        self.destination_folder_path_var = tk.StringVar()

        # --- Widgets for Lookup File ---
        tk.Label(master, text="Product Lookup File (.xlsx):").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        tk.Entry(master, textvariable=self.lookup_file_path_var, width=60).grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        tk.Button(master, text="Browse", command=self.browse_lookup_file).grid(row=0, column=2, padx=10, pady=5)

        # --- Widgets for Source Folder ---
        tk.Label(master, text="Source Folder (Master Files):").grid(row=1, column=0, sticky="w", padx=10, pady=5)
        tk.Entry(master, textvariable=self.source_folder_path_var, width=60).grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        tk.Button(master, text="Browse", command=self.browse_source_folder).grid(row=1, column=2, padx=10, pady=5)

        # --- Widgets for Destination Folder ---
        tk.Label(master, text="Destination Folder (New Files):").grid(row=2, column=0, sticky="w", padx=10, pady=5)
        tk.Entry(master, textvariable=self.destination_folder_path_var, width=60).grid(row=2, column=1, padx=10, pady=5, sticky="ew")
        tk.Button(master, text="Browse", command=self.browse_destination_folder).grid(row=2, column=2, padx=10, pady=5)

        # --- Execute Button ---
        tk.Button(master, text="Run Duplication & Highlight", command=self.run_duplication,
                  font=("Arial", 12, "bold"), bg="#4CAF50", fg="white").grid(row=3, column=0, columnspan=3, pady=20)

        # --- Status / Log Area ---
        tk.Label(master, text="Process Log:").grid(row=4, column=0, sticky="w", padx=10, pady=5)
        self.log_text = scrolledtext.ScrolledText(master, wrap=tk.WORD, width=80, height=15, font=("Courier New", 10))
        self.log_text.grid(row=5, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")
        # Make the log area resizable
        master.grid_rowconfigure(5, weight=1)

    def log_message(self, message, color="black"):
        """Appends a message to the log area."""
        self.log_text.insert(tk.END, message + "\n", color)
        self.log_text.see(tk.END) # Scroll to the end
        # Define tags for colors
        self.log_text.tag_config("red", foreground="red")
        self.log_text.tag_config("green", foreground="green")
        self.log_text.tag_config("blue", foreground="blue")
        self.log_text.tag_config("orange", foreground="orange")
        self.master.update_idletasks() # Update GUI immediately

    def browse_lookup_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Product Lookup Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.lookup_file_path_var.set(file_path)
            self.log_message(f"Selected lookup file: {file_path}", "blue")

    def browse_source_folder(self):
        folder_path = filedialog.askdirectory(title="Select Folder with Master Excel Files")
        if folder_path:
            self.source_folder_path_var.set(folder_path)
            self.log_message(f"Selected source folder: {folder_path}", "blue")

    def browse_destination_folder(self):
        folder_path = filedialog.askdirectory(title="Select Destination Folder for New Files")
        if folder_path:
            self.destination_folder_path_var.set(folder_path)
            self.log_message(f"Selected destination folder: {folder_path}", "blue")

    def run_duplication(self):
        self.log_text.delete(1.0, tk.END) # Clear previous log
        self.log_message("Starting Excel file duplication process...", "blue")

        lookup_file = self.lookup_file_path_var.get()
        source_folder = self.source_folder_path_var.get()
        destination_folder = self.destination_folder_path_var.get()

        # Basic validation
        if not lookup_file or not os.path.exists(lookup_file):
            self.log_message("Error: Please select a valid product lookup Excel file.", "red")
            messagebox.showerror("Input Error", "Please select a valid product lookup Excel file.")
            return
        if not source_folder or not os.path.isdir(source_folder):
            self.log_message("Error: Please select a valid source folder.", "red")
            messagebox.showerror("Input Error", "Please select a valid source folder.")
            return
        if not destination_folder:
            self.log_message("Error: Please select a destination folder.", "red")
            messagebox.showerror("Input Error", "Please select a destination folder.")
            return

        # Create destination folder if it doesn't exist
        if not os.path.exists(destination_folder):
            try:
                os.makedirs(destination_folder)
                self.log_message(f"Created destination folder: {destination_folder}")
            except Exception as e:
                self.log_message(f"Error creating destination folder '{destination_folder}': {e}", "red")
                messagebox.showerror("Folder Creation Error", f"Could not create destination folder:\n{e}")
                return

        # Read the lookup list using pandas for processing logic
        try:
            lookup_df = pd.read_excel(lookup_file)
            self.log_message(f"Successfully loaded lookup list with {len(lookup_df)} entries.", "blue")
            if "Product Code" not in lookup_df.columns or "Product Name" not in lookup_df.columns:
                self.log_message("Error: Lookup file must contain 'Product Code' and 'Product Name' columns.", "red")
                self.log_message(f"Available columns: {lookup_df.columns.tolist()}", "red")
                messagebox.showerror("Lookup File Error", "The lookup file must have 'Product Code' and 'Product Name' columns.")
                return
        except FileNotFoundError:
            self.log_message(f"Error: Lookup file not found at '{lookup_file}'.", "red")
            messagebox.showerror("File Not Found", f"Lookup file not found:\n{lookup_file}")
            return
        except Exception as e:
            self.log_message(f"Error reading lookup file '{lookup_file}': {e}", "red")
            messagebox.showerror("File Read Error", f"Error reading lookup file:\n{e}")
            return

        processed_count = 0
        skipped_count = 0
        # Set to store product codes that were not found in the source folder
        not_found_product_codes = set()

        # Iterate through the lookup list
        for index, row in lookup_df.iterrows():
            product_code = str(row["Product Code"]).strip()
            product_name = str(row["Product Name"]).strip()

            source_file_name = f"{product_name}.xlsx"
            source_file_path = os.path.join(source_folder, source_file_name)
            destination_file_name = f"{product_code}.xlsx"
            destination_file_path = os.path.join(destination_folder, destination_file_name)

            self.log_message(f"Processing '{product_code}' - '{product_name}':")
            self.log_message(f"  Source file expected: {source_file_name}")

            if os.path.exists(source_file_path):
                try:
                    shutil.copy2(source_file_path, destination_file_path)
                    self.log_message(f"  -> Created '{destination_file_name}'", "green")
                    processed_count += 1
                except Exception as e:
                    self.log_message(f"  Error copying file: {e}", "red")
                    skipped_count += 1
                    not_found_product_codes.add(product_code) # Add to skipped list for highlighting
            else:
                self.log_message(f"  Source file NOT FOUND: '{source_file_name}'. Skipping.", "orange")
                skipped_count += 1
                not_found_product_codes.add(product_code) # Add to skipped list for highlighting

        self.log_message("\n--- Duplication Process Complete ---", "blue")
        self.log_message(f"Total entries processed: {processed_count}", "blue")
        self.log_message(f"Total entries skipped: {skipped_count}", "blue")

        # --- Highlighting Logic ---
        if not_found_product_codes:
            self.log_message(f"\nHighlighting {len(not_found_product_codes)} skipped product codes in lookup file...", "blue")
            try:
                wb = load_workbook(lookup_file)
                ws = wb.active # Get the active worksheet (usually the first one)

                # Find the column index for "Product Code"
                product_code_col_idx = -1
                for col_idx, cell in enumerate(ws[1]): # Iterate through the first row (headers)
                    if cell.value == "Product Code":
                        product_code_col_idx = col_idx + 1 # openpyxl is 1-indexed for columns
                        break

                if product_code_col_idx == -1:
                    self.log_message("Warning: 'Product Code' column not found in lookup file for highlighting.", "orange")
                else:
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    # Iterate through rows starting from the second row (skipping header)
                    for row_idx in range(2, ws.max_row + 1):
                        cell_value = ws.cell(row=row_idx, column=product_code_col_idx).value
                        if str(cell_value).strip() in not_found_product_codes:
                            ws.cell(row=row_idx, column=product_code_col_idx).fill = yellow_fill
                            self.log_message(f"  Highlighted '{cell_value}' in lookup file.", "green")
                    wb.save(lookup_file)
                    self.log_message(f"Lookup file '{lookup_file}' updated with highlighting.", "blue")

            except Exception as e:
                self.log_message(f"Error during highlighting in lookup file: {e}", "red")
                messagebox.showerror("Highlighting Error", f"Could not highlight product codes in lookup file:\n{e}")
        else:
            self.log_message("No product codes needed highlighting (all source files found or no items skipped).", "blue")

        messagebox.showinfo("Process Complete",
                            f"Duplication process finished!\n"
                            f"Processed: {processed_count}\n"
                            f"Skipped: {skipped_count}\n"
                            f"Files saved to: {destination_folder}\n"
                            f"Lookup file '{os.path.basename(lookup_file)}' has been updated with highlighting for skipped items.")

# Main part to run the GUI application
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelDuplicatorApp(root)
    root.mainloop()
